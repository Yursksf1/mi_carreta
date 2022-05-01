from django.shortcuts import render

# Create your views here.
from django.views.generic.edit import FormView
from datetime import date, timedelta
from tzlocal import get_localzone

from django.shortcuts import redirect, render
from django.views.generic.edit import CreateView, DeleteView, UpdateView
from django.urls import reverse
from django.db.models import Q
import datetime

from myapp.models import HistoryWeight, HistoryWeather, SheepBreed, Sheep, HistoryWeight, Observations, HistoryPluviometer, HistoryBodyCondition, HistoryFamacha, Group, Breed
from myapp.serializers import HistoryWeatherSerialize
from myapp.controllers import SheepController
from django.http import FileResponse
from django.db.models.functions import TruncMonth

IMAGEN_LOCATION = {
    'N': 'img/nevera.png',
    'L': 'img/laboratorio.png',
    'A': 'img/aprisco.png',
}

def weather_today(request):
    weathers_today = []

    for locations in HistoryWeather.LOCATION:
        current_weather = HistoryWeather.objects.filter(
            location=locations[0],
            create_at__day=date.today().day,
            create_at__year=date.today().year,
            create_at__month=date.today().month,
        ).first()

        serializer = HistoryWeatherSerialize(current_weather)
        current_weather = serializer.data
        current_weather['location'] = locations[1]
        current_weather['image_location'] = IMAGEN_LOCATION.get(locations[0])
        weathers_today.append(current_weather)


    print('weathers_today', weathers_today)
    return render(
        request,
        'weather_today.html',
        {
            'weathers_today': weathers_today,
        }
    )


def weather_history(request):
    nevera_date = []
    nevera_data_t = []
    nevera_data_h = []

    laboratorio_date = []
    laboratorio_data_t = []
    laboratorio_data_h = []

    aprisco_date = []
    aprisco_data_t = []
    aprisco_data_h = []

    query = HistoryWeather.objects.filter(
        location='N'
    )
    if query.exists():
        results = query.order_by('-create_at').all()[:100]
        for result in results:
            nevera_date.append(str(result.create_at.date()))
            nevera_data_t.append(float(result.temperature))
            nevera_data_h.append(float(result.humidity))

    query = HistoryWeather.objects.filter(
        location='L'
    )
    if query.exists():
        results = query.order_by('-create_at').all()[:100]
        for result in results:
            laboratorio_date.append(str(result.create_at.date()))
            laboratorio_data_t.append(float(result.temperature))
            laboratorio_data_h.append(float(result.humidity))

    query = HistoryWeather.objects.filter(
        location='A'
    )
    if query.exists():
        results = query.order_by('-create_at').all()[:100]
        for result in results:
            aprisco_date.append(str(result.create_at.date()))
            aprisco_data_t.append(float(result.temperature))
            aprisco_data_h.append(float(result.humidity))


    return render(
        request,
        'weather_history.html',
        {
            'nevera_date': nevera_date,
            'nevera_data_t': nevera_data_t,
            'nevera_data_h': nevera_data_h,

            'laboratorio_date': laboratorio_date,
            'laboratorio_data_t': laboratorio_data_t,
            'laboratorio_data_h': laboratorio_data_h,

            'aprisco_date': aprisco_date,
            'aprisco_data_t': aprisco_data_t,
            'aprisco_data_h': aprisco_data_h,
        }
    )

def dashboard(request):
    sheeps = Sheep.objects
    ids_pure = SheepController.get_id_pure()
    today = date.today()
    month_6 = today - timedelta(days=180)
    month_12 = today - timedelta(days=360)

    total = len(sheeps.filter(active=True).all())

    num_machos_p = len(sheeps.filter(gender='M', id__in=ids_pure, active=True))
    num_hembras_p = len(sheeps.filter(gender='H', id__in=ids_pure, active=True))
    num_machos_c = len(sheeps.filter(gender='M', active=True).exclude(id__in=ids_pure))
    num_hembras_c = len(sheeps.filter(gender='H', active=True).exclude(id__in=ids_pure))

    num_ovejas = len(sheeps.filter(birthday__lt=month_12, active=True))
    num_borregos = len(sheeps.filter(birthday__lt=month_6, birthday__gt=month_12,  active=True))
    num_corderos = len(sheeps.filter(birthday__gt=month_6, active=True))

    num_ovejas_m_p = len(sheeps.filter(birthday__lt=month_12, gender='M', id__in=ids_pure, active=True))
    num_ovejas_h_p = len(sheeps.filter(birthday__lt=month_12, gender='H', id__in=ids_pure, active=True))
    num_borregos_m_p = len(sheeps.filter(birthday__lt=month_6, birthday__gt=month_12, gender='M', id__in=ids_pure, active=True))
    num_borregos_h_p = len(sheeps.filter(birthday__lt=month_6, birthday__gt=month_12, gender='H', id__in=ids_pure, active=True))
    num_corderos_m_p = len(sheeps.filter(birthday__gt=month_6, gender='M', id__in=ids_pure, active=True))
    num_corderos_h_p = len(sheeps.filter(birthday__gt=month_6, gender='H', id__in=ids_pure, active=True))

    num_ovejas_m_c = len(sheeps.filter(birthday__lt=month_12, gender='M', active=True).exclude(id__in=ids_pure))
    num_ovejas_h_c = len(sheeps.filter(birthday__lt=month_12, gender='H', active=True).exclude(id__in=ids_pure))
    num_borregos_m_c = len(sheeps.filter(birthday__lt=month_6, birthday__gt=month_12, gender='M', active=True).exclude(id__in=ids_pure))
    num_borregos_h_c = len(sheeps.filter(birthday__lt=month_6, birthday__gt=month_12, gender='H', active=True).exclude(id__in=ids_pure))
    num_corderos_m_c = len(sheeps.filter(birthday__gt=month_6, gender='M', active=True).exclude(id__in=ids_pure))
    num_corderos_h_c = len(sheeps.filter(birthday__gt=month_6, gender='H', active=True).exclude(id__in=ids_pure))

    num_gestantes = 11
    num_natal = 8
    num_destetes = 20

    num_vendidas = 25
    num_muertes = 6
    groups = Group.objects.all()

    context = {
            'total': total,

            'num_machos_p': num_machos_p,
            'num_hembras_p': num_hembras_p,
            'num_machos_c': num_machos_c,
            'num_hembras_c': num_hembras_c,

            'num_ovejas': num_ovejas,
            'num_borregos': num_borregos,
            'num_corderos': num_corderos,

            'num_ovejas_m_p': num_ovejas_m_p,
            'num_ovejas_h_p': num_ovejas_h_p,
            'num_borregos_m_p': num_borregos_m_p,
            'num_borregos_h_p': num_borregos_h_p,
            'num_corderos_m_p': num_corderos_m_p,
            'num_corderos_h_p': num_corderos_h_p,

            'num_ovejas_m_c': num_ovejas_m_c,
            'num_ovejas_h_c': num_ovejas_h_c,
            'num_borregos_m_c': num_borregos_m_c,
            'num_borregos_h_c': num_borregos_h_c,
            'num_corderos_m_c': num_corderos_m_c,
            'num_corderos_h_c': num_corderos_h_c,

            # aun falta buscar esto:
            'num_gestantes': num_gestantes,
            'num_natal': num_natal,
            'num_destetes': num_destetes,
            'num_vendidas': num_vendidas,
            'num_muertes': num_muertes,

            'groups': groups

        }

    return render(
        request,
        'dashboard.html',
        context
    )

class WeatherCreateView(CreateView):
    template_name = 'weather_form.html'
    model = HistoryWeather
    fields = ['humidity', 'temperature', 'location']

    def get_success_url(self):
        return reverse('app:weather-today')


# SHEEPS
# Django
from django.urls import reverse_lazy
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import CreateView, DetailView, ListView

# Forms
from .forms import SheepForm

# Models
from .models import Sheep

class SheepsFeedView(ListView):
    """Return all published sheeps."""

    template_name = 'sheep_list.html'
    model = Sheep
    paginate_by = 200
    context_object_name = 'sheeps'

    def get_queryset(self):
        """Filter by price if it is provided in GET parameters"""
        queryset = super().get_queryset()

        if 'active' in self.request.GET:
            if self.request.GET['active'] == 'true':
                queryset = queryset.filter(active=True)
        if 'name' in self.request.GET:
            queryset = queryset.filter(name__icontains=self.request.GET['name'])
        if 'search' in self.request.GET:
            queryset = queryset.filter(
                    Q(identification_number=self.request.GET['search']) |
                    Q(identification_number_2=self.request.GET['search']) |
                    Q(name__icontains=self.request.GET['search'])
                )
        if 'gender' in self.request.GET:
            queryset = queryset.filter(gender=self.request.GET['gender'])

        if 'breeds' in self.request.GET:
            ids_pure = SheepController.get_id_pure()

            if self.request.GET['breeds'] == 'puro':
                queryset = queryset.filter(
                    id__in=ids_pure
                )

            if self.request.GET['breeds'] == 'comercial':
                queryset = queryset.exclude(id__in=ids_pure)

        if 'group' in self.request.GET:
            group_name = self.request.GET['group']
            ids_in_group = SheepController.get_id_in_group(group_name)
            queryset = queryset.filter(
                id__in=ids_in_group
            )

        if 'type' in self.request.GET:
            if self.request.GET['type'] == 'cordero':
                today = date.today()
                month_6 = today - timedelta(days=180)
                queryset = queryset.filter(
                    birthday__gt=month_6
                )

            if self.request.GET['type'] == 'borrego':
                today = date.today()
                month_6 = today - timedelta(days=180)
                month_12 = today - timedelta(days=360)
                queryset = queryset.filter(
                    birthday__lt=month_6,
                    birthday__gt=month_12
                )

            if self.request.GET['type'] == 'oveja':
                today = date.today()
                month_12 = today - timedelta(days=360)
                queryset = queryset.filter(
                    birthday__lt=month_12
                )


        if 'age_range_max' in self.request.GET:
            age_range_max = self.request.GET['age_range_max']
            if age_range_max.isnumeric():
                age_range_max = int(age_range_max)
                today = date.today()
                age_range_day_max = today - timedelta(days=age_range_max)
                queryset = queryset.filter(birthday__gt=age_range_day_max)

        if 'age_range_min' in self.request.GET:
            age_range_min = self.request.GET['age_range_min']
            if age_range_min.isnumeric():
                age_range_min = int(age_range_min)
                today = date.today()
                age_range_day_min = today - timedelta(days=age_range_min)
                queryset = queryset.filter(birthday__lt=age_range_day_min)

        if 'weight_range_max' in self.request.GET or 'weight_range_min' in self.request.GET:

            weight_range_max = None
            if 'weight_range_max' in self.request.GET:
                weight_range_max = self.request.GET['weight_range_max']
                if weight_range_max.isnumeric():
                    weight_range_max = int(weight_range_max)
                else:
                    weight_range_max = None

            weight_range_min = None
            if 'weight_range_min' in self.request.GET:
                weight_range_min = self.request.GET['weight_range_min']
                if weight_range_min.isnumeric():
                    weight_range_min = int(weight_range_min)
                else:
                    weight_range_min = None

            ids_range_weight = SheepController.get_id_range_weight(weight_range_min, weight_range_max)

            queryset = queryset.filter(
                id__in=ids_range_weight
            )

        if True:
            queryset = queryset.order_by('birthday')

        return queryset



class SheepDetailView(DetailView):
    """Return sheep detail."""

    template_name = 'sheep_detail.html'
    queryset = Sheep.objects.all()
    context_object_name = 'sheep'

class SheepDetailHistoryView(DetailView):
    """Return sheep detail."""

    template_name = 'sheep_detail_history.html'
    queryset = Sheep.objects.all()
    context_object_name = 'sheep'


class CreateSheepView(CreateView):
    """Create a new sheep."""

    template_name = 'sheep_form.html'
    form_class = SheepForm
    success_url = reverse_lazy('app:feed')

    def get_context_data(self, **kwargs):
        """Add user and profile to context."""
        context = super().get_context_data(**kwargs)
        context['user'] = self.request.user
        return context

from django.views import View
from django.shortcuts import get_object_or_404

class SheepWeighView(View):

    def get(self, request, pk, *args, **kwargs):
        sheep = get_object_or_404(Sheep, pk=pk)
        context = {
            'sheep': sheep
        }

        return render(
            request,
            'sheep_weigh_form.html',
            context
        )

    def post(self, request, pk, *args, **kwargs):
        data_request = request.POST
        if data_request:
            data = {
                'date': data_request.get('date'),
                'weight': data_request.get('weight'),
                'famacha': data_request.get('famacha'),
                'corporal': data_request.get('corporal'),
                'evaluation-reproductive': data_request.get('evaluation-reproductive'),
                'observacion': data_request.get('observacion'),
            }
            SheepController().update_weight(pk, data)
        return redirect('app:detail', pk)


class ObservationsView(ListView):
    """Return all published sheeps."""

    template_name = 'observations.html'
    model = Observations
    paginate_by = 20
    context_object_name = 'observation'

    def get_queryset(self):
        """Filter by price if it is provided in GET parameters"""
        queryset = super().get_queryset()

        queryset = queryset.filter(active=True)
        queryset = queryset.order_by('-create_at')

        return queryset


def observations_check(request, pk):
    observation = Observations.objects.filter(pk=pk).first()
    if observation:
        observation.active = False
        observation.save()
    return redirect('app:observations')

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

def generate_document(hoja, datas):
    start = 1
    years = datas.keys()
    for year in years:
        generate_year(hoja, start, year, datas.get(year))
        start = start + 13

def generate_year(hoja, start, year, year_data):
    hoja.merge_cells("A{}:AF{}".format(start, start))
    hoja = add_yead(hoja, start, year)
    hoja = generate_months(hoja, start, year_data)

    return hoja

def add_yead(hoja, start, year):
    position = "A{}".format(start)
    celda = hoja[position]
    hoja[position] = year
    skyblue = PatternFill(start_color='87CEEB',
                          end_color='87CEEB',
                          fill_type='solid')

    celda.fill = skyblue
    celda.font = Font(name='Times New Roman', size=18, color='000000', bold=True)
    celda.alignment = Alignment(horizontal='center')

    return hoja


def generate_months(hoja, start, year_data):
    encabezado = ["MESES", "1", "2", "3", "4", "5", "6", "7", "8", "9", "10", "11", "12", "13", "14", "15", "16", "17", "18", "19", "20", "21", "22", "23", "24", "25", "26", "27", "28", "29", "30", "31", "TOTAL MES ", "PROMEDIO MES"]
    hoja.append(encabezado)
    for index, mes in enumerate(meses):
        if index+1 in year_data.keys():
            registro = [mes] + year_data.get(index+1) + ['=AVERAGE(B{i}:AF{i})'.format(i=hoja.max_row +1), '=SUM(B{i}:AF{i})'.format(i=hoja.max_row+1)]
            hoja.append(registro)

    return hoja

def prepare_data_pluviometer(data_pluviometer):
    data = {
    }
    for pluviometer in data_pluviometer:
        year = pluviometer.create_at.year
        month = pluviometer.create_at.month
        day = pluviometer.create_at.day - 1
        if not data.get(year):
            data[year] = {}
        if not data.get(year).get(month):
            data[year][month] = [0] * 31
        data[year][month][day] = pluviometer.measure

    return data

columnas = [
    "B",
    "C",
    "D",
    "E",
    "F",
    "G",
    "H",
    "I",
    "J",
    "K",
    "L",
    "M",
    "N",
    "O",
    "P",
    "Q",
    "R",
    "S",
    "T",
    "U",
    "V",
    "W",
    "X",
    "Y",
    "Z",
    "AA",
    "AB",
    "AC",
    "AD",
    "AE",
    "AF",
]

meses = [
    'ENERO',
    'FEBRERO',
    'MARZO',
    'ABRIL',
    'MAYO',
    'JUNIO',
    'JULIO',
    'AGOSTO',
    'SEPTIEMBRE',
    'OCTUBRE',
    'NOVIEMBRE',
    'DICIEMBRE',
]

def pluviometer_download(request):
    path = 'media/generate_reports'
    file_name = 'pluvimetro.xlsx'
    file_path = '{}/{}'.format(path, file_name)
    libro = Workbook()
    hoja = libro.active
    for i in columnas:
        hoja.column_dimensions[i].width = 5

    pluviometer_data = HistoryPluviometer.objects.all()
    pluviometer_data = prepare_data_pluviometer(pluviometer_data)

    hoja = generate_document(hoja, pluviometer_data)

    libro.save(file_path)

    response = FileResponse(open(file_path, 'rb'))
    return response

def pluviometer_import(request):
    excel_file = request.FILES["excel_file"]

    # you may put validations here to check extension or file size

    wb = openpyxl.load_workbook(excel_file)

    # getting a particular sheet by name out of many sheets
    worksheet = wb["Sheet"]
    print(worksheet)

    excel_data = list()
    # iterating over the rows and
    # getting value from each cell in row
    year = 2020
    month = "ENERO"
    for row in worksheet.iter_rows():
        row_data = list()
        valor_a = row[0].value
        if type(valor_a) == int:
            print("es del año: {}".format(valor_a))
            year = valor_a
        elif valor_a in meses:
            month = meses.index(valor_a) + 1

            for index, cell in enumerate(row):
                if index == 0:
                    continue

                new_value = cell.value
                if not new_value:
                    continue
                if index in range(1, 31):
                    day = index
                    print("{} {} {} value: {}".format(day, month, year, new_value))

                    date = datetime.date(year, month, day)
                    hp = HistoryPluviometer.objects.filter(create_at=date).first()
                    if hp:
                        print('update HistoryPluviometer')
                        hp.measure = new_value
                    else:
                        print('create HistoryPluviometer')
                        hp = HistoryPluviometer()
                        hp.create_at = date
                        hp.measure = new_value
                    hp.save()


    return redirect('app:acciones_bloque')

## EJEMPLO
def pluviometer_download_2(request):
    path = 'media/generate_reports'
    file_name = 'pluvimetro.xlsx'
    file_path = '{}/{}'.format(path, file_name)

    libro = Workbook()
    hoja = libro.active
    celda = hoja["B2"]
    hoja["B2"] = "Muestra"
    celda.font = Font(name='Arial', size=14, color='00FF0000')

    libro.save(file_path)

    response = FileResponse(open(file_path, 'rb'))
    return response

## weather

def generate_document_weather(hoja, weather_data):
    encabezado = ["FECHA", "UBICACION", "TEMPERATURA", "HUMEDAD"]
    hoja.append(encabezado)
    LOCATION = {
        "N": 'Nevera',
        "L": 'Laboratorio',
        "A": 'Aprisco',
    }
    for wd in weather_data:
        data = [date.isoformat(wd.create_at), LOCATION.get(wd.location), wd.temperature, wd.humidity]
        hoja.append(data)

def weather_download(request):
    path = 'media/generate_reports'
    file_name = 'weather.xlsx'
    file_path = '{}/{}'.format(path, file_name)
    libro = Workbook()
    hoja = libro.active
    for i in columnas:
        hoja.column_dimensions[i].width = 10

    weather_data = HistoryWeather.objects.order_by('create_at').all()

    hoja = generate_document_weather(hoja, weather_data)

    libro.save(file_path)

    response = FileResponse(open(file_path, 'rb'))
    return response

def weather_import(request):
    excel_file = request.FILES["excel_file"]

    # you may put validations here to check extension or file size

    wb = openpyxl.load_workbook(excel_file)

    # getting a particular sheet by name out of many sheets
    worksheet = wb["Sheet"]
    print(worksheet)

    for index, row in enumerate(worksheet.iter_rows()):
        if index == 0:
            continue

        date = row[0].value
        location = row[1].value
        temperature = row[2].value
        humidity = row[3].value

        date = datetime.date.fromisoformat(date)
        hw = HistoryWeather.objects.filter(create_at=date, location=location).first()
        if hw:
            print('update HistoryPluviometer')
            hw.temperature = temperature
            hw.humidity = humidity
        else:
            print('create HistoryPluviometer')
            hw = HistoryWeather()
            hw.create_at = date
            hw.location = location
            hw.temperature = temperature
            hw.humidity = humidity
        hw.save()
    return redirect('app:acciones_bloque')


## Weight


def generate_document_weight(hoja, sheeps_data):
    encabezado = ["Nombre", "#OVINO", "GENERO", "Edad", "ultimo peso", "ultima famacha", "ultima condicion corportal"]
    hoja.append(encabezado)

    for sheep in sheeps_data:
        data = [sheep.name, sheep.identification_number, sheep.gender, sheep.age(), sheep.last_weight(), sheep.last_famacha(), sheep.last_body_condition()]
        hoja.append(data)

def weight_download(request):
    path = 'media/generate_reports'
    file_name = 'weight.xlsx'
    file_path = '{}/{}'.format(path, file_name)
    libro = Workbook()
    hoja = libro.active

    sheeps = Sheep.objects.filter(active=True).order_by("active", "-gender", "identification_number").all()
    # here is the magic
    generate_document_weight(hoja, sheeps)

    libro.save(file_path)
    response = FileResponse(open(file_path, 'rb'))
    return response

def weight_import(request):
    excel_file = request.FILES["excel_file"]

    # you may put validations here to check extension or file size

    wb = openpyxl.load_workbook(excel_file)

    # getting a particular sheet by name out of many sheets
    worksheet = wb["Sheet"]
    print(worksheet)
    index_ovino = 1
    index_weight = 4
    index_famacha = 5
    index_body_condition = 6

    for index, row in enumerate(worksheet.iter_rows()):
        if index == 0:
            continue
        else:
            num_ovino = row[index_ovino].value
            sheep = Sheep.objects.filter(identification_number=num_ovino).first()
            if not sheep:
                print("no se encuentra registro de {}".format(num_ovino))
                continue

            weight = row[index_weight].value
            famacha = row[index_famacha].value
            body_condition = row[index_body_condition].value

            local_tz = get_localzone()
            today = datetime.datetime.now(local_tz).date()
            if weight:
                hw = HistoryWeight.objects.filter(create_at__date=today, sheep=sheep).first()
                if hw:
                    print('update HistoryPluviometer')
                    hw.weight = weight
                else:
                    print('create HistoryPluviometer')
                    hw = HistoryWeight()
                    hw.weight = weight
                    hw.sheep = sheep
                hw.save()
            if body_condition:

                hbc = HistoryBodyCondition.objects.filter(create_at__date=today, sheep=sheep).first()
                if hbc:
                    print('update HistoryPluviometer')
                    hbc.body_condition = body_condition
                else:
                    print('create HistoryPluviometer')
                    hbc = HistoryBodyCondition()
                    hbc.body_condition = body_condition
                    hbc.sheep = sheep
                hbc.save()

            if famacha:
                hf = HistoryFamacha.objects.filter(create_at__date=today, sheep=sheep).first()
                if hf:
                    print('update HistoryPluviometer')
                    hf.famacha = famacha
                else:
                    print('create HistoryPluviometer')
                    hf = HistoryFamacha()
                    hf.famacha = famacha
                    hf.sheep = sheep
                hf.save()

    return redirect('app:acciones_bloque')


# multiple weight

def generate_document_weights(hoja, sheeps_data):
    encabezado = ["Nombre", "#OVINO", "GENERO", "Edad"]

    hw = HistoryWeight.objects.values('create_at').group_by('create_at')
    hoja.append(encabezado)

    for sheep in sheeps_data:
        data = [sheep.name, sheep.identification_number, sheep.gender, sheep.age(), sheep.last_weight(), sheep.last_famacha(), sheep.last_body_condition()]
        hoja.append(data)

def weights_download(request):
    path = 'media/generate_reports'
    file_name = 'multiple_weight.xlsx'
    file_path = '{}/{}'.format(path, file_name)
    libro = Workbook()
    hoja = libro.active

    sheeps = Sheep.objects.filter(active=True).order_by("active", "-gender", "identification_number").all()
    # here is the magic
    generate_document_weights(hoja, sheeps)

    libro.save(file_path)
    response = FileResponse(open(file_path, 'rb'))
    return response

def weights_import(request):
    excel_file = request.FILES["excel_file"]

    # you may put validations here to check extension or file size

    wb = openpyxl.load_workbook(excel_file)

    # getting a particular sheet by name out of many sheets
    worksheet = wb["Sheet"]
    print(worksheet)
    index_ovino = 1
    index_weight = 4
    index_famacha = 5
    index_body_condition = 6

    for index, row in enumerate(worksheet.iter_rows()):
        if index == 0:
            continue
        else:
            num_ovino = row[index_ovino].value
            sheep = Sheep.objects.filter(identification_number=num_ovino).first()
            if not sheep:
                print("no se encuentra registro de {}".format(num_ovino))
                continue

            weight = row[index_weight].value
            famacha = row[index_famacha].value
            body_condition = row[index_body_condition].value

            local_tz = get_localzone()
            today = datetime.datetime.now(local_tz).date()
            if weight:
                hw = HistoryWeight.objects.filter(create_at__date=today, sheep=sheep).first()
                if hw:
                    print('update HistoryPluviometer')
                    hw.weight = weight
                else:
                    print('create HistoryPluviometer')
                    hw = HistoryWeight()
                    hw.weight = weight
                    hw.sheep = sheep
                hw.save()
            if body_condition:

                hbc = HistoryBodyCondition.objects.filter(create_at__date=today, sheep=sheep).first()
                if hbc:
                    print('update HistoryPluviometer')
                    hbc.body_condition = body_condition
                else:
                    print('create HistoryPluviometer')
                    hbc = HistoryBodyCondition()
                    hbc.body_condition = body_condition
                    hbc.sheep = sheep
                hbc.save()

            if famacha:
                hf = HistoryFamacha.objects.filter(create_at__date=today, sheep=sheep).first()
                if hf:
                    print('update HistoryPluviometer')
                    hf.famacha = famacha
                else:
                    print('create HistoryPluviometer')
                    hf = HistoryFamacha()
                    hf.famacha = famacha
                    hf.sheep = sheep
                hf.save()

    return redirect('app:acciones_bloque')


# multiple sheep

def generate_document_sheep(hoja, sheeps_data):
    encabezado = ["Activo",  "Nombre", "#OVINO", "GENERO", "nacimiento", "padre", "madre"]

    hoja.append(encabezado)

    for sheep in sheeps_data:
        active = "N"
        if sheep.active:
            active = "S"
        data = [active, sheep.name, sheep.identification_number, sheep.gender, sheep.birthday.strftime("%d/%m/%Y")]
        parents = ["", ""]
        if sheep.parentDadId:
            parents[0] = sheep.parentDadId.identification_number
        if sheep.parentMomId:
            parents[1] = sheep.parentMomId.identification_number
        data = data + parents
        hoja.append(data)

def sheep_download(request):
    path = 'media/generate_reports'
    file_name = 'sheep_import.xlsx'
    file_path = '{}/{}'.format(path, file_name)
    libro = Workbook()
    hoja = libro.active

    sheeps = Sheep.objects.order_by("active", "-gender", "identification_number").all()
    # here is the magic
    generate_document_sheep(hoja, sheeps)

    libro.save(file_path)
    response = FileResponse(open(file_path, 'rb'))
    return response

def sheep_import(request):
    excel_file = request.FILES["excel_file"]

    # you may put validations here to check extension or file size

    wb = openpyxl.load_workbook(excel_file)

    # getting a particular sheet by name out of many sheets
    worksheet = wb["Sheet"]
    print(worksheet)
    index_is_active = 0
    index_name = 1
    index_num = 2
    index_gender = 3
    index_birthday = 4
    index_dad = 5
    index_mom = 6

    for index, row in enumerate(worksheet.iter_rows()):
        if index == 0:
            continue
        else:
            is_active = row[index_is_active].value
            num_ovino = row[index_num].value
            name = row[index_name].value
            gender = row[index_gender].value
            date_birthday = row[index_birthday].value
            local_tz = get_localzone()
            birthday = datetime.datetime.now(local_tz).date()
            if date_birthday:
                birthday = datetime.datetime.strptime(date_birthday,  '%d/%m/%Y')
            dad = row[index_dad].value
            mom = row[index_mom].value
            sheep = Sheep.objects.filter(identification_number=num_ovino).first()
            if not sheep:
                sheep = Sheep()
                sheep.name = name
                sheep.identification_number = num_ovino
                sheep.gender = gender
                sheep.birthday = birthday
                sheep.active = not (is_active.upper() == 'N')
                sheep.save()

                if dad:
                    sheep_dad = Sheep.objects.filter(identification_number=dad).first()
                    if sheep_dad:
                        sheep.parentDadId = sheep_dad

                if mom:
                    sheep_mom = Sheep.objects.filter(identification_number=mom).first()
                    if sheep_mom:
                        sheep.parentMomId = sheep_mom
                sheep.save()




    return redirect('app:acciones_bloque')


# multiple sheep breeds

def generate_document_sheep_breeds(hoja, sheeps_data):
    encabezado = ["oveja_id",  "rasa_acronim", "percentage"]

    hoja.append(encabezado)

    for sheep in sheeps_data:
        breeds = sheep.breed.all()
        for breed in breeds:
            if breed != "N":
                data = [sheep.identification_number, breed.breed.acronym, int(breed.percent)]
                hoja.append(data)

def sheep_breeds_download(request):
    path = 'media/generate_reports'
    file_name = 'sheep_breeds_import.xlsx'
    file_path = '{}/{}'.format(path, file_name)
    libro = Workbook()
    hoja = libro.active

    sheeps = Sheep.objects.order_by("active", "-gender", "identification_number").all()
    # here is the magic
    generate_document_sheep_breeds(hoja, sheeps)

    libro.save(file_path)
    response = FileResponse(open(file_path, 'rb'))
    return response

def sheep_breeds_import(request):
    excel_file = request.FILES["excel_file"]

    # you may put validations here to check extension or file size

    wb = openpyxl.load_workbook(excel_file)

    # getting a particular sheet by name out of many sheets
    worksheet = wb["Sheet"]
    print(worksheet)

    index_ov_num = 0
    index_acr = 1
    index_percentaje = 2

    for index, row in enumerate(worksheet.iter_rows()):
        if index == 0:
            continue
        else:
            num_ovino = row[index_ov_num].value
            acronym = row[index_acr].value
            percetaje = row[index_percentaje].value

            sheep = Sheep.objects.filter(identification_number=num_ovino).first()
            breed = Breed.objects.filter(acronym=acronym).first()
            if sheep and breed:
                sheep_breed = SheepBreed.objects.filter(sheep=sheep, breed=breed, percetaje=percetaje)
                if not sheep_breed:
                    sheep_breed = SheepBreed()
                    sheep_breed.sheep = sheep
                    sheep_breed.breed = breed
                    sheep_breed.percetaje = percetaje
                    sheep_breed.save()

    return redirect('app:acciones_bloque')
